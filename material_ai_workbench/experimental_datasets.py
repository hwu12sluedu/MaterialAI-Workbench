"""Governed ingestion for public experimental composite datasets.

The source workbook is never bundled with the package.  Users either provide a
locally downloaded copy or explicitly accept the source license before the
workbench downloads it into the configured workspace.
"""

from __future__ import annotations

import csv
import hashlib
import json
import math
import os
import shutil
import tempfile
import urllib.request
import warnings
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable
from uuid import uuid4

from material_ai_workbench.config import DATASETS_ROOT


@dataclass(frozen=True)
class ExperimentalDatasetSpec:
    dataset_id: str
    version: str
    title: str
    filename: str
    file_id: str
    content_id: str
    sha256: str
    size_bytes: int
    article_doi: str
    dataset_doi: str
    article_url: str
    dataset_url: str
    metadata_url: str
    download_url: str
    license_name: str
    license_url: str


ALSHEGHRI_CFRP_SPEC = ExperimentalDatasetSpec(
    dataset_id="alsheghri_2025_cfrp_experiment",
    version="1",
    title=(
        "Predicting mechanical properties of CFRP composites using "
        "data-driven models with comparative analysis"
    ),
    filename="Dataset_Processing-Structure-Property_CFRPs.xlsx",
    file_id="b3ddf6cd-0827-4ed1-a7df-c23955da9c91",
    content_id="5b823a27-7797-4b8b-be12-e8d2edda5cfd",
    sha256="759fe8568ac38a7d45db0aaa65b4aa3d8d9386d46615411e530d375d5f6652e8",
    size_bytes=15745,
    article_doi="10.1371/journal.pone.0319787",
    dataset_doi="10.17632/fspdwb4mst.1",
    article_url=(
        "https://journals.plos.org/plosone/article?" "id=10.1371/journal.pone.0319787"
    ),
    dataset_url="https://data.mendeley.com/datasets/fspdwb4mst/1",
    metadata_url=(
        "https://data.mendeley.com/public-api/datasets/fspdwb4mst/files?"
        "folder_id=root&version=1"
    ),
    download_url=(
        "https://data.mendeley.com/public-files/datasets/fspdwb4mst/files/"
        "b3ddf6cd-0827-4ed1-a7df-c23955da9c91/file_downloaded"
    ),
    license_name="CC BY-NC 3.0",
    license_url="https://creativecommons.org/licenses/by-nc/3.0/",
)


SOURCE_HEADERS = (
    "PROPERTIES",
    "Carbon Nanotube Volume Fraction",
    "Interlayer Volume Fraction (EPFOAM/ELSP)",
    "Glass Transition Temperature (\u1d52C)",
    "Manufacturing Pressure (psi)",
    "Flexural Strength (MPa)",
    "Flexural Modulus (MPa)",
    "Tension Strength (MPa)",
    "Mode II energy-release rate kJ/m2 (Interlayer Improvement)",
)

NORMALIZED_COLUMNS = (
    "sample_id",
    "source_row",
    "material_type_id",
    "material_type_name",
    "cnt_volume_fraction",
    "interlayer_volume_fraction",
    "glass_transition_temperature_c",
    "manufacturing_pressure_psi",
    "flexural_strength_mpa",
    "flexural_modulus_reported_mpa",
    "tensile_strength_mpa",
    "mode_ii_energy_release_rate_reported_kj_m2",
)

FEATURE_COLUMNS = (
    "cnt_volume_fraction",
    "interlayer_volume_fraction",
    "glass_transition_temperature_c",
    "manufacturing_pressure_psi",
)

TARGET_COLUMNS = (
    "flexural_strength_mpa",
    "flexural_modulus_reported_mpa",
    "tensile_strength_mpa",
    "mode_ii_energy_release_rate_reported_kj_m2",
)

MATERIAL_TYPES = {
    1: "Control CFRP 1",
    2: "Epoxy Foam CFRP",
    3: "CNT Epoxy Foam CFRP",
    4: "ELSP CFRP",
    5: "CNT ELSP CFRP",
    6: "Pressure MFG - 30 psi",
    7: "Pressure MFG - 50 psi",
    8: "Pressure MFG - 70 psi",
    9: "Control CFRP 2",
}

EXPECTED_GROUP_COUNTS = {1: 14, 2: 8, 3: 6, 4: 7, 5: 7, 6: 4, 7: 5, 8: 5, 9: 6}

COLUMN_SCHEMA = (
    {"name": "sample_id", "role": "identifier", "unit": None, "source": "derived"},
    {"name": "source_row", "role": "provenance", "unit": None, "source": "derived"},
    {
        "name": "material_type_id",
        "role": "group",
        "unit": None,
        "source": SOURCE_HEADERS[0],
    },
    {
        "name": "material_type_name",
        "role": "group_label",
        "unit": None,
        "source": "derived from source legend",
    },
    {
        "name": "cnt_volume_fraction",
        "role": "feature",
        "unit": "dimensionless",
        "source": SOURCE_HEADERS[1],
    },
    {
        "name": "interlayer_volume_fraction",
        "role": "feature",
        "unit": "dimensionless",
        "source": SOURCE_HEADERS[2],
    },
    {
        "name": "glass_transition_temperature_c",
        "role": "feature",
        "unit": "degC",
        "source": SOURCE_HEADERS[3],
    },
    {
        "name": "manufacturing_pressure_psi",
        "role": "feature",
        "unit": "psi",
        "source": SOURCE_HEADERS[4],
    },
    {
        "name": "flexural_strength_mpa",
        "role": "target",
        "unit": "MPa",
        "source": SOURCE_HEADERS[5],
    },
    {
        "name": "flexural_modulus_reported_mpa",
        "role": "target",
        "unit": "MPa (source label, unverified)",
        "source": SOURCE_HEADERS[6],
    },
    {
        "name": "tensile_strength_mpa",
        "role": "target",
        "unit": "MPa",
        "source": SOURCE_HEADERS[7],
    },
    {
        "name": "mode_ii_energy_release_rate_reported_kj_m2",
        "role": "target",
        "unit": "kJ/m2 (source label, unverified)",
        "source": SOURCE_HEADERS[8],
    },
)


@dataclass(frozen=True)
class ExperimentalDatasetResult:
    dataset_dir: Path
    source_workbook: Path
    normalized_csv: Path
    manifest_json: Path
    quality_report_json: Path
    split_manifest_json: Path
    data_card_md: Path
    row_count: int
    quality_status: str


def prepare_cfrp_experimental_dataset(
    *,
    source_path: Path | str | None = None,
    output_root: Path | str = DATASETS_ROOT,
    accept_license: bool = False,
    timeout_seconds: float = 60.0,
    spec: ExperimentalDatasetSpec = ALSHEGHRI_CFRP_SPEC,
) -> ExperimentalDatasetResult:
    """Verify, normalize, audit, and split the public CFRP workbook.

    ``accept_license`` is required only when the workbench downloads the file.
    A caller that supplies ``source_path`` is responsible for having obtained
    that file under the source dataset's license.
    """

    root = Path(output_root).expanduser().resolve()
    dataset_dir = root / spec.dataset_id / f"v{spec.version}"
    raw_dir = dataset_dir / "raw"
    derived_dir = dataset_dir / "derived"
    raw_dir.mkdir(parents=True, exist_ok=True)
    derived_dir.mkdir(parents=True, exist_ok=True)

    temporary_download: Path | None = None
    obtained_via: str
    original_source: str
    try:
        if source_path is None:
            if not accept_license:
                raise ValueError(
                    f"Downloading this dataset requires explicit acceptance of "
                    f"{spec.license_name}; pass accept_license=True."
                )
            metadata = fetch_mendeley_file_metadata(
                spec, timeout_seconds=timeout_seconds
            )
            download_url = str(metadata["download_url"])
            temporary_download = _download_to_temporary_file(
                download_url,
                parent=dataset_dir,
                timeout_seconds=timeout_seconds,
            )
            candidate = temporary_download
            obtained_via = "official_download"
            original_source = download_url
        else:
            candidate = Path(source_path).expanduser().resolve()
            if not candidate.is_file():
                raise FileNotFoundError(f"Source workbook does not exist: {candidate}")
            obtained_via = "local_file"
            original_source = candidate.name

        _verify_source_file(candidate, spec)
        source_workbook = raw_dir / spec.filename
        _install_verified_source(candidate, source_workbook, spec)
    finally:
        if temporary_download is not None:
            temporary_download.unlink(missing_ok=True)

    rows = read_cfrp_experimental_workbook(source_workbook)
    quality = build_cfrp_quality_report(rows, spec=spec)
    splits = build_grouped_split_manifest(rows, spec=spec)

    normalized_csv = derived_dir / "cfrp_experimental_normalized.csv"
    quality_report_json = derived_dir / "quality_report.json"
    split_manifest_json = derived_dir / "grouped_splits.json"
    data_card_md = dataset_dir / "DATA_CARD_CN.md"
    manifest_json = dataset_dir / "dataset_manifest.json"

    _write_rows_csv(normalized_csv, rows)
    _write_json(quality_report_json, quality)
    _write_json(split_manifest_json, splits)
    _write_text(data_card_md, _render_data_card(spec, rows, quality, splits))

    manifest = {
        "schema_version": 1,
        "created_at": _utc_now(),
        "dataset_id": spec.dataset_id,
        "dataset_version": spec.version,
        "title": spec.title,
        "source": {
            "article_doi": spec.article_doi,
            "dataset_doi": spec.dataset_doi,
            "article_url": spec.article_url,
            "dataset_url": spec.dataset_url,
            "license": spec.license_name,
            "license_url": spec.license_url,
            "obtained_via": obtained_via,
            "original_source": original_source,
            "filename": spec.filename,
            "sha256": _sha256(source_workbook),
            "size_bytes": source_workbook.stat().st_size,
        },
        "row_count": len(rows),
        "features": list(FEATURE_COLUMNS),
        "targets": list(TARGET_COLUMNS),
        "columns": [dict(column) for column in COLUMN_SCHEMA],
        "group_column": "material_type_id",
        "quality_status": quality["status"],
        "artifacts": {
            "source_workbook": _artifact_entry(source_workbook, dataset_dir),
            "normalized_csv": _artifact_entry(normalized_csv, dataset_dir),
            "quality_report": _artifact_entry(quality_report_json, dataset_dir),
            "grouped_splits": _artifact_entry(split_manifest_json, dataset_dir),
            "data_card": _artifact_entry(data_card_md, dataset_dir),
        },
        "limitations": [
            "This is a small experimental tabular dataset, not a constitutive dataset.",
            "The source unit labels are retained without silent conversion.",
            "Rows with unavailable targets are retained and excluded per target fold.",
            "Exact duplicate rows are retained for source fidelity.",
        ],
    }
    _write_json(manifest_json, manifest)

    return ExperimentalDatasetResult(
        dataset_dir=dataset_dir,
        source_workbook=source_workbook,
        normalized_csv=normalized_csv,
        manifest_json=manifest_json,
        quality_report_json=quality_report_json,
        split_manifest_json=split_manifest_json,
        data_card_md=data_card_md,
        row_count=len(rows),
        quality_status=str(quality["status"]),
    )


def fetch_mendeley_file_metadata(
    spec: ExperimentalDatasetSpec = ALSHEGHRI_CFRP_SPEC,
    *,
    timeout_seconds: float = 30.0,
) -> dict[str, Any]:
    """Read and validate the current official file metadata."""

    request = urllib.request.Request(
        spec.metadata_url,
        headers={
            "Accept": "application/vnd.mendeley-public-dataset.1+json",
            "User-Agent": "MaterialAI-Workbench/0.4",
        },
    )
    with urllib.request.urlopen(request, timeout=timeout_seconds) as response:
        payload = json.loads(response.read().decode("utf-8"))
    return validate_mendeley_file_metadata(payload, spec=spec)


def validate_mendeley_file_metadata(
    payload: Any,
    *,
    spec: ExperimentalDatasetSpec = ALSHEGHRI_CFRP_SPEC,
) -> dict[str, Any]:
    """Validate Mendeley file-list JSON without downloading the workbook."""

    if isinstance(payload, dict):
        entries = payload.get("files") or payload.get("data") or payload.get("value")
    else:
        entries = payload
    if not isinstance(entries, list):
        raise ValueError("Mendeley metadata response does not contain a file list.")

    match = next(
        (
            item
            for item in entries
            if isinstance(item, dict)
            and (
                item.get("id") == spec.file_id or item.get("filename") == spec.filename
            )
        ),
        None,
    )
    if match is None:
        raise ValueError(f"Official dataset file not found: {spec.filename}")

    details = match.get("content_details")
    if not isinstance(details, dict):
        details = {}
    actual_hash = str(
        details.get("sha256_hash")
        or match.get("sha256")
        or match.get("sha256_hash")
        or match.get("file_hash")
        or ""
    ).lower()
    actual_size = details.get("size") or match.get("size") or match.get("size_bytes")
    content_id = details.get("id") or match.get("content_id") or match.get("contentId")
    if not actual_hash:
        raise ValueError("Official dataset metadata does not include SHA-256.")
    if actual_hash != spec.sha256.lower():
        raise ValueError(
            "Official dataset SHA-256 changed; review the new version first."
        )
    if actual_size is None:
        raise ValueError("Official dataset metadata does not include file size.")
    if int(actual_size) != spec.size_bytes:
        raise ValueError("Official dataset size changed; review the new version first.")
    if content_id is None:
        raise ValueError("Official dataset metadata does not include content id.")
    if str(content_id) != spec.content_id:
        raise ValueError(
            "Official dataset content id changed; review the new version first."
        )
    normalized = dict(match)
    normalized.update(
        {
            "sha256": actual_hash,
            "size_bytes": int(actual_size),
            "content_id": str(content_id),
            "download_url": details.get("download_url") or spec.download_url,
        }
    )
    return normalized


def read_cfrp_experimental_workbook(path: Path | str) -> list[dict[str, Any]]:
    """Parse the known source workbook into a stable, unit-labelled schema."""

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is packaged
        raise RuntimeError(
            "openpyxl is required for CFRP workbook ingestion. Reinstall the package."
        ) from exc

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Unknown extension is not supported and will be removed",
            module="openpyxl.worksheet._reader",
        )
        workbook = load_workbook(Path(path), read_only=True, data_only=True)
        try:
            return _parse_cfrp_workbook(workbook)
        finally:
            workbook.close()


def _parse_cfrp_workbook(workbook: Any) -> list[dict[str, Any]]:
    if "Dataset" not in workbook.sheetnames:
        raise ValueError("Expected worksheet 'Dataset' was not found.")
    worksheet = workbook["Dataset"]
    actual_headers = tuple(
        _clean_header(cell.value) for cell in worksheet[11][: len(SOURCE_HEADERS)]
    )
    if actual_headers != SOURCE_HEADERS:
        raise ValueError(
            "Unexpected CFRP workbook headers. "
            f"Expected {SOURCE_HEADERS!r}, got {actual_headers!r}."
        )

    rows: list[dict[str, Any]] = []
    within_group: Counter[int] = Counter()
    for source_row, values in enumerate(
        worksheet.iter_rows(
            min_row=13,
            max_col=len(SOURCE_HEADERS),
            values_only=True,
        ),
        start=13,
    ):
        if all(value is None for value in values):
            continue
        material_type_id = _required_int(values[0], source_row, SOURCE_HEADERS[0])
        if material_type_id not in MATERIAL_TYPES:
            raise ValueError(
                f"Row {source_row}: unsupported material type {material_type_id}."
            )
        within_group[material_type_id] += 1
        rows.append(
            {
                "sample_id": (
                    f"cfrp_type_{material_type_id:02d}_"
                    f"sample_{within_group[material_type_id]:02d}"
                ),
                "source_row": source_row,
                "material_type_id": material_type_id,
                "material_type_name": MATERIAL_TYPES[material_type_id],
                "cnt_volume_fraction": _required_float(
                    values[1], source_row, SOURCE_HEADERS[1]
                ),
                "interlayer_volume_fraction": _required_float(
                    values[2], source_row, SOURCE_HEADERS[2]
                ),
                "glass_transition_temperature_c": _required_float(
                    values[3], source_row, SOURCE_HEADERS[3]
                ),
                "manufacturing_pressure_psi": _required_float(
                    values[4], source_row, SOURCE_HEADERS[4]
                ),
                "flexural_strength_mpa": _required_float(
                    values[5], source_row, SOURCE_HEADERS[5]
                ),
                "flexural_modulus_reported_mpa": _required_float(
                    values[6], source_row, SOURCE_HEADERS[6]
                ),
                "tensile_strength_mpa": _optional_float(
                    values[7], source_row, SOURCE_HEADERS[7]
                ),
                "mode_ii_energy_release_rate_reported_kj_m2": _optional_float(
                    values[8], source_row, SOURCE_HEADERS[8]
                ),
            }
        )

    if len(rows) != 62:
        raise ValueError(f"Expected 62 CFRP samples, found {len(rows)}.")
    counts = Counter(int(row["material_type_id"]) for row in rows)
    if dict(sorted(counts.items())) != EXPECTED_GROUP_COUNTS:
        raise ValueError(
            "Unexpected material type counts: "
            f"expected {EXPECTED_GROUP_COUNTS}, got {dict(sorted(counts.items()))}."
        )
    return rows


def build_cfrp_quality_report(
    rows: list[dict[str, Any]],
    *,
    spec: ExperimentalDatasetSpec = ALSHEGHRI_CFRP_SPEC,
) -> dict[str, Any]:
    """Build an auditable profile without dropping or imputing source rows."""

    missing = {
        column: sum(row[column] is None for row in rows)
        for column in FEATURE_COLUMNS + TARGET_COLUMNS
    }
    ranges = {
        column: _numeric_profile(row[column] for row in rows)
        for column in FEATURE_COLUMNS + TARGET_COLUMNS
    }
    group_counts = Counter(int(row["material_type_id"]) for row in rows)
    signature_counts = Counter(
        tuple(row[column] for column in NORMALIZED_COLUMNS[2:]) for row in rows
    )
    duplicate_extra_rows = sum(
        count - 1 for count in signature_counts.values() if count > 1
    )
    duplicate_record_count = sum(
        count for count in signature_counts.values() if count > 1
    )

    warnings = [
        "Only 62 samples are available; uncertainty and grouped validation are mandatory.",
        "The source labels flexural modulus as MPa; values are retained without conversion.",
        "The source labels mode-II energy release rate as kJ/m2; values are retained without conversion.",
    ]
    if duplicate_extra_rows:
        warnings.append(
            f"{duplicate_extra_rows} exact duplicate rows beyond first occurrences are retained."
        )
    for target in TARGET_COLUMNS:
        if missing[target]:
            warnings.append(
                f"{target} is unavailable for {missing[target]} rows; no imputation was applied."
            )

    return {
        "schema_version": 1,
        "created_at": _utc_now(),
        "dataset_id": spec.dataset_id,
        "status": "pass_with_warnings" if warnings else "pass",
        "row_count": len(rows),
        "column_count": len(NORMALIZED_COLUMNS),
        "group_counts": {str(key): group_counts[key] for key in sorted(group_counts)},
        "missing_counts": missing,
        "numeric_profiles": ranges,
        "duplicate_extra_rows": duplicate_extra_rows,
        "duplicate_record_count": duplicate_record_count,
        "checks": {
            "expected_row_count": len(rows) == 62,
            "expected_group_counts": dict(sorted(group_counts.items()))
            == EXPECTED_GROUP_COUNTS,
            "features_complete": all(
                missing[column] == 0 for column in FEATURE_COLUMNS
            ),
            "required_targets_complete": all(
                missing[column] == 0 for column in TARGET_COLUMNS[:2]
            ),
        },
        "warnings": warnings,
        "transformations": [
            "Replaced source 'Not available' target cells with empty values.",
            "Added stable sample ids, source row ids, and material type names.",
            "No unit conversion, imputation, outlier removal, or deduplication was applied.",
        ],
    }


def build_grouped_split_manifest(
    rows: list[dict[str, Any]],
    *,
    spec: ExperimentalDatasetSpec = ALSHEGHRI_CFRP_SPEC,
) -> dict[str, Any]:
    """Create target-specific leave-one-material-type-out evaluation folds."""

    targets: dict[str, Any] = {}
    for target in TARGET_COLUMNS:
        available = [row for row in rows if row[target] is not None]
        groups = sorted({int(row["material_type_id"]) for row in available})
        folds: list[dict[str, Any]] = []
        for test_group in groups:
            train_ids = [
                str(row["sample_id"])
                for row in available
                if int(row["material_type_id"]) != test_group
            ]
            test_ids = [
                str(row["sample_id"])
                for row in available
                if int(row["material_type_id"]) == test_group
            ]
            folds.append(
                {
                    "fold_id": f"holdout_type_{test_group:02d}",
                    "test_group": test_group,
                    "test_group_name": MATERIAL_TYPES[test_group],
                    "train_sample_ids": train_ids,
                    "test_sample_ids": test_ids,
                    "train_count": len(train_ids),
                    "test_count": len(test_ids),
                }
            )
        targets[target] = {
            "available_sample_count": len(available),
            "available_groups": groups,
            "fold_count": len(folds),
            "folds": folds,
        }

    return {
        "schema_version": 1,
        "created_at": _utc_now(),
        "dataset_id": spec.dataset_id,
        "strategy": "leave_one_material_type_out",
        "group_column": "material_type_id",
        "purpose": (
            "Prevent samples from the same CFRP type appearing in both training "
            "and evaluation data."
        ),
        "targets": targets,
    }


def _verify_source_file(path: Path, spec: ExperimentalDatasetSpec) -> None:
    size = path.stat().st_size
    digest = _sha256(path)
    if size != spec.size_bytes:
        raise ValueError(
            f"Source size mismatch: expected {spec.size_bytes}, found {size}."
        )
    if digest.lower() != spec.sha256.lower():
        raise ValueError(
            f"Source SHA-256 mismatch: expected {spec.sha256}, found {digest}."
        )


def _install_verified_source(
    candidate: Path,
    destination: Path,
    spec: ExperimentalDatasetSpec,
) -> None:
    if destination.exists():
        _verify_source_file(destination, spec)
        return
    temporary = destination.with_name(f".{destination.name}.{uuid4().hex}.partial")
    try:
        shutil.copy2(candidate, temporary)
        _verify_source_file(temporary, spec)
        os.replace(temporary, destination)
    finally:
        temporary.unlink(missing_ok=True)


def _download_to_temporary_file(
    url: str,
    *,
    parent: Path,
    timeout_seconds: float,
) -> Path:
    descriptor, name = tempfile.mkstemp(
        prefix="cfrp-source-", suffix=".xlsx", dir=parent
    )
    os.close(descriptor)
    path = Path(name)
    request = urllib.request.Request(
        url,
        headers={"User-Agent": "MaterialAI-Workbench/0.4"},
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout_seconds) as response:
            with path.open("wb") as handle:
                shutil.copyfileobj(response, handle)
    except Exception:
        path.unlink(missing_ok=True)
        raise
    return path


def _clean_header(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _required_int(value: Any, row: int, column: str) -> int:
    number = _required_float(value, row, column)
    if not float(number).is_integer():
        raise ValueError(f"Row {row}, {column}: expected an integer, got {value!r}.")
    return int(number)


def _required_float(value: Any, row: int, column: str) -> float:
    result = _optional_float(value, row, column)
    if result is None:
        raise ValueError(f"Row {row}, {column}: a numeric value is required.")
    return result


def _optional_float(value: Any, row: int, column: str) -> float | None:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if text.lower() in {"not available", "n/a", "na"}:
            return None
        text = text.replace(",", "")
    else:
        text = value
    try:
        number = float(text)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"Row {row}, {column}: invalid numeric value {value!r}."
        ) from exc
    if not math.isfinite(number):
        raise ValueError(f"Row {row}, {column}: value must be finite.")
    return number


def _numeric_profile(values: Iterable[Any]) -> dict[str, Any]:
    numeric = [float(value) for value in values if value is not None]
    if not numeric:
        return {"count": 0, "min": None, "max": None, "mean": None}
    return {
        "count": len(numeric),
        "min": min(numeric),
        "max": max(numeric),
        "mean": sum(numeric) / len(numeric),
    }


def _write_rows_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.{uuid4().hex}.tmp")
    try:
        with temporary.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=list(NORMALIZED_COLUMNS))
            writer.writeheader()
            for row in rows:
                writer.writerow(
                    {
                        key: "" if row[key] is None else row[key]
                        for key in NORMALIZED_COLUMNS
                    }
                )
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _write_json(path: Path, payload: Any) -> None:
    _write_text(path, json.dumps(payload, ensure_ascii=False, indent=2) + "\n")


def _write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.{uuid4().hex}.tmp")
    try:
        temporary.write_text(content, encoding="utf-8")
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _artifact_entry(path: Path, root: Path) -> dict[str, Any]:
    return {
        "path": _relative(path, root),
        "sha256": _sha256(path),
        "size_bytes": path.stat().st_size,
    }


def _relative(path: Path, root: Path) -> str:
    return path.relative_to(root).as_posix()


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _render_data_card(
    spec: ExperimentalDatasetSpec,
    rows: list[dict[str, Any]],
    quality: dict[str, Any],
    splits: dict[str, Any],
) -> str:
    missing = quality["missing_counts"]
    groups = quality["group_counts"]
    group_lines = "\n".join(
        f"- 类型 `{key}`（{MATERIAL_TYPES[int(key)]}）：{count} 条"
        for key, count in groups.items()
    )
    target_lines = "\n".join(
        f"- `{target}`：可用 {details['available_sample_count']} 条，"
        f"{details['fold_count']} 个按类型留一验证折"
        for target, details in splits["targets"].items()
    )
    warning_lines = "\n".join(f"- {warning}" for warning in quality["warnings"])
    return f"""# CFRP 公开实验数据卡

## 身份与许可

- 数据集：{spec.title}
- 论文 DOI：`{spec.article_doi}`
- 数据 DOI：`{spec.dataset_doi}`（版本 `{spec.version}`）
- 许可：[{spec.license_name}]({spec.license_url})，仅按许可允许的范围使用
- 原始文件 SHA-256：`{spec.sha256}`
- 标准化样本数：`{len(rows)}`

原始 Excel 不随 MaterialAI Workbench 或 GitHub 仓库分发。下载前必须显式接受
源数据许可；工作区中的 `dataset_manifest.json` 保存来源、哈希和派生文件信息。

## 这个数据能回答什么

它适合研究“加工与夹层参数 -> CFRP 宏观性能”的小样本回归，包括弯曲强度、
源文件所标注的弯曲模量、拉伸强度和 II 型能量释放率。它不能直接训练 Abaqus
材料本构，因为没有完整应力-应变历史、铺层细节、损伤演化参数或微观场数据。

## 规范化字段

输入特征：`{', '.join(FEATURE_COLUMNS)}`。

预测目标：`{', '.join(TARGET_COLUMNS)}`。

分组字段：`material_type_id`，用于阻止同一 CFRP 类型同时进入训练集和测试集。

## 样本分组

{group_lines}

## 缺失值

- 拉伸强度缺失：`{missing['tensile_strength_mpa']}` 条
- II 型能量释放率缺失：`{missing['mode_ii_energy_release_rate_reported_kj_m2']}` 条
- 输入特征缺失：`0` 条

缺失目标保留为空值；训练每个目标时只使用该目标有效的样本，不进行自动填补。

## 可信划分

采用 `leave_one_material_type_out`。每折完整留出一种材料类型，而不是随机拆分行：

{target_lines}

## 已知风险

{warning_lines}

源文件把弯曲模量标为 MPa、把 II 型能量释放率标为 kJ/m2。本工具使用
`reported` 字段名保留原值和原单位标签，不做静默换算。将数据映射到真实 Abaqus
模型前，必须回查试验定义和论文。

## 建议用途

1. 先用 Ridge、Random Forest 和 SVR 建立论文对照基线。
2. 用按材料类型留一验证报告均值、方差和每一折误差，不能只报一次随机切分。
3. 再加入 MLP 等模型；若小样本下没有稳定提升，应保留传统模型。
4. 把本数据作为实验锚点，而不是把 62 行表格包装成“有限元神经网络”。
"""


__all__ = [
    "ALSHEGHRI_CFRP_SPEC",
    "COLUMN_SCHEMA",
    "ExperimentalDatasetResult",
    "ExperimentalDatasetSpec",
    "FEATURE_COLUMNS",
    "TARGET_COLUMNS",
    "build_cfrp_quality_report",
    "build_grouped_split_manifest",
    "fetch_mendeley_file_metadata",
    "prepare_cfrp_experimental_dataset",
    "read_cfrp_experimental_workbook",
    "validate_mendeley_file_metadata",
]
