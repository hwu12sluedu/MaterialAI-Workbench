from __future__ import annotations

import csv
import hashlib
import json
from dataclasses import replace
from pathlib import Path

import pytest
from openpyxl import Workbook

from material_ai_workbench.experimental_datasets import (
    ALSHEGHRI_CFRP_SPEC,
    EXPECTED_GROUP_COUNTS,
    SOURCE_HEADERS,
    prepare_cfrp_experimental_dataset,
    validate_mendeley_file_metadata,
)
from material_ai_workbench.run_experimental_dataset import main as dataset_cli_main


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _write_source_workbook(path: Path, *, bad_header: bool = False) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dataset"
    sheet.cell(row=1, column=1, value="Type of Samples")
    headers = list(SOURCE_HEADERS)
    if bad_header:
        headers[1] = "Unexpected CNT field"
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=11, column=column, value=header)

    source_row = 13
    for material_type, count in EXPECTED_GROUP_COUNTS.items():
        for index in range(count):
            pressure = {6: 30, 7: 50, 8: 70}.get(material_type, 0)
            interlayer = (
                0.08
                if material_type in {2, 3}
                else 0.066 if material_type in {4, 5} else 0
            )
            cnt = 0.0136 if material_type in {3, 5} else 0
            values = [
                material_type,
                cnt,
                interlayer,
                50 + material_type * 10 + index * 0.1,
                pressure,
                180 + material_type * 20 + index,
                20 + material_type * 3 + index * 0.2,
                230 + material_type + index if material_type <= 5 else "Not available",
                (
                    300 + material_type * 50 + index
                    if material_type <= 5
                    else "Not available"
                ),
            ]
            for column, value in enumerate(values, start=1):
                sheet.cell(row=source_row, column=column, value=value)
            source_row += 1
    workbook.save(path)


def _fixture_spec(path: Path):
    return replace(
        ALSHEGHRI_CFRP_SPEC,
        sha256=_sha256(path),
        size_bytes=path.stat().st_size,
    )


def test_prepare_cfrp_dataset_writes_traceable_artifacts(tmp_path: Path) -> None:
    source = tmp_path / ALSHEGHRI_CFRP_SPEC.filename
    _write_source_workbook(source)
    spec = _fixture_spec(source)

    result = prepare_cfrp_experimental_dataset(
        source_path=source,
        output_root=tmp_path / "datasets",
        spec=spec,
    )

    assert result.row_count == 62
    assert result.quality_status == "pass_with_warnings"
    assert result.source_workbook.read_bytes() == source.read_bytes()
    assert result.manifest_json.exists()
    assert result.quality_report_json.exists()
    assert result.split_manifest_json.exists()
    assert result.data_card_md.exists()

    with result.normalized_csv.open("r", encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))
    assert len(rows) == 62
    assert rows[0]["sample_id"] == "cfrp_type_01_sample_01"
    assert rows[-1]["tensile_strength_mpa"] == ""

    manifest = json.loads(result.manifest_json.read_text(encoding="utf-8"))
    assert manifest["source"]["sha256"] == spec.sha256
    assert manifest["source"]["license"] == "CC BY-NC 3.0"
    assert manifest["source"]["obtained_via"] == "local_file"
    assert manifest["source"]["original_source"] == source.name
    assert manifest["quality_status"] == "pass_with_warnings"
    columns = {column["name"]: column for column in manifest["columns"]}
    assert columns["material_type_id"]["role"] == "group"
    assert columns["flexural_modulus_reported_mpa"]["unit"].endswith(
        "(source label, unverified)"
    )

    quality = json.loads(result.quality_report_json.read_text(encoding="utf-8"))
    assert quality["checks"]["features_complete"] is True
    assert quality["missing_counts"]["tensile_strength_mpa"] == 20
    assert quality["missing_counts"]["mode_ii_energy_release_rate_reported_kj_m2"] == 20


def test_grouped_folds_never_mix_test_group_into_training(tmp_path: Path) -> None:
    source = tmp_path / ALSHEGHRI_CFRP_SPEC.filename
    _write_source_workbook(source)
    result = prepare_cfrp_experimental_dataset(
        source_path=source,
        output_root=tmp_path / "datasets",
        spec=_fixture_spec(source),
    )

    splits = json.loads(result.split_manifest_json.read_text(encoding="utf-8"))
    assert splits["strategy"] == "leave_one_material_type_out"
    assert splits["targets"]["flexural_strength_mpa"]["fold_count"] == 9
    assert splits["targets"]["tensile_strength_mpa"]["fold_count"] == 5

    for details in splits["targets"].values():
        for fold in details["folds"]:
            train = set(fold["train_sample_ids"])
            test = set(fold["test_sample_ids"])
            assert train.isdisjoint(test)
            prefix = f"cfrp_type_{fold['test_group']:02d}_"
            assert test
            assert all(sample_id.startswith(prefix) for sample_id in test)
            assert all(not sample_id.startswith(prefix) for sample_id in train)


def test_prepare_rejects_hash_mismatch_before_install(tmp_path: Path) -> None:
    source = tmp_path / ALSHEGHRI_CFRP_SPEC.filename
    _write_source_workbook(source)
    wrong_hash_spec = replace(
        ALSHEGHRI_CFRP_SPEC,
        size_bytes=source.stat().st_size,
    )

    with pytest.raises(ValueError, match="SHA-256 mismatch"):
        prepare_cfrp_experimental_dataset(
            source_path=source,
            output_root=tmp_path / "datasets",
            spec=wrong_hash_spec,
        )

    assert not (
        tmp_path
        / "datasets"
        / ALSHEGHRI_CFRP_SPEC.dataset_id
        / "v1"
        / "raw"
        / ALSHEGHRI_CFRP_SPEC.filename
    ).exists()


def test_prepare_rejects_unexpected_workbook_schema(tmp_path: Path) -> None:
    source = tmp_path / ALSHEGHRI_CFRP_SPEC.filename
    _write_source_workbook(source, bad_header=True)

    with pytest.raises(ValueError, match="Unexpected CFRP workbook headers"):
        prepare_cfrp_experimental_dataset(
            source_path=source,
            output_root=tmp_path / "datasets",
            spec=_fixture_spec(source),
        )


def test_download_requires_explicit_license_acceptance(tmp_path: Path) -> None:
    with pytest.raises(ValueError, match="explicit acceptance"):
        prepare_cfrp_experimental_dataset(output_root=tmp_path / "datasets")


def test_cli_returns_structured_error_for_missing_source(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    exit_code = dataset_cli_main(
        [
            "--source-xlsx",
            str(tmp_path / "missing.xlsx"),
            "--output-root",
            str(tmp_path / "datasets"),
        ]
    )

    captured = capsys.readouterr()
    assert exit_code == 2
    assert captured.out == ""
    payload = json.loads(captured.err)
    assert payload["status"] == "error"
    assert "does not exist" in payload["error"]


def test_mendeley_metadata_validation_checks_nested_content_details() -> None:
    spec = ALSHEGHRI_CFRP_SPEC
    payload = [
        {
            "filename": spec.filename,
            "id": spec.file_id,
            "size": spec.size_bytes,
            "content_details": {
                "id": spec.content_id,
                "sha256_hash": spec.sha256,
                "size": spec.size_bytes,
                "download_url": spec.download_url,
            },
        }
    ]

    metadata = validate_mendeley_file_metadata(payload)

    assert metadata["sha256"] == spec.sha256
    assert metadata["content_id"] == spec.content_id
    assert metadata["download_url"] == spec.download_url

    payload[0]["content_details"]["sha256_hash"] = "0" * 64
    with pytest.raises(ValueError, match="SHA-256 changed"):
        validate_mendeley_file_metadata(payload)
