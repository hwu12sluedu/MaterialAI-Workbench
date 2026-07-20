from __future__ import annotations

import csv
import hashlib
import json
import pickle
import shutil
from collections import Counter
from dataclasses import replace
from pathlib import Path

import pytest
from openpyxl import Workbook

from material_ai_workbench.experimental_baselines import (
    load_cfrp_baseline_dataset,
    train_cfrp_grouped_baselines,
    validate_target_split_contract,
)
from material_ai_workbench.experimental_datasets import (
    ALSHEGHRI_CFRP_SPEC,
    EXPECTED_GROUP_COUNTS,
    SOURCE_HEADERS,
    prepare_cfrp_experimental_dataset,
)
from material_ai_workbench.experimental_validation import run_cfrp_validation_audit
from material_ai_workbench.run_experimental_baselines import main as baseline_cli_main
from material_ai_workbench.run_experimental_validation import (
    main as validation_audit_cli_main,
)


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _write_source_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dataset"
    for column, header in enumerate(SOURCE_HEADERS, start=1):
        sheet.cell(row=11, column=column, value=header)

    source_row = 13
    for material_type, count in EXPECTED_GROUP_COUNTS.items():
        for index in range(count):
            pressure = {6: 30, 7: 50, 8: 70}.get(material_type, 0)
            interlayer = (
                0.08
                if material_type in {2, 3}
                else 0.066 if material_type in {4, 5} else 0
            )
            cnt = 0.0136 if material_type in {3, 5} else 0
            values = [
                material_type,
                cnt,
                interlayer,
                50 + material_type * 10 + index * 0.1,
                pressure,
                180 + material_type * 20 + index,
                20 + material_type * 3 + index * 0.2,
                230 + material_type + index if material_type <= 5 else "Not available",
                (
                    300 + material_type * 50 + index
                    if material_type <= 5
                    else "Not available"
                ),
            ]
            for column, value in enumerate(values, start=1):
                sheet.cell(row=source_row, column=column, value=value)
            source_row += 1
    workbook.save(path)


def _prepare_dataset(root: Path) -> Path:
    source = root / ALSHEGHRI_CFRP_SPEC.filename
    source.parent.mkdir(parents=True, exist_ok=True)
    _write_source_workbook(source)
    spec = replace(
        ALSHEGHRI_CFRP_SPEC,
        sha256=_sha256(source),
        size_bytes=source.stat().st_size,
    )
    result = prepare_cfrp_experimental_dataset(
        source_path=source,
        output_root=root / "datasets",
        spec=spec,
    )
    return result.dataset_dir


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def _refresh_manifest_artifact(dataset_dir: Path, artifact_name: str) -> None:
    manifest_path = dataset_dir / "dataset_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    entry = manifest["artifacts"][artifact_name]
    artifact = dataset_dir / entry["path"]
    entry["sha256"] = _sha256(artifact)
    entry["size_bytes"] = artifact.stat().st_size
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def _make_exact_duplicate(
    dataset_dir: Path, *, retained_sample_id: str, duplicate_sample_id: str
) -> None:
    normalized = dataset_dir / "derived" / "cfrp_experimental_normalized.csv"
    rows = _read_csv(normalized)
    by_id = {row["sample_id"]: row for row in rows}
    retained = by_id[retained_sample_id]
    duplicate = by_id[duplicate_sample_id]
    for column in rows[0]:
        if column not in {"sample_id", "source_row"}:
            duplicate[column] = retained[column]
    with normalized.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)
    _refresh_manifest_artifact(dataset_dir, "normalized_csv")


def test_grouped_baselines_write_oof_metrics_models_and_figures(tmp_path: Path) -> None:
    dataset_dir = _prepare_dataset(tmp_path / "source")

    run = train_cfrp_grouped_baselines(
        dataset_dir,
        output_root=tmp_path / "experiments",
        targets=("flexural_strength_mpa", "tensile_strength_mpa"),
        models=("mean", "ridge", "random_forest", "svr"),
        random_state=7,
        rf_estimators=12,
    )

    assert run.manifest_json.exists()
    assert run.summary_json.exists()
    assert run.comparison_csv.exists()
    assert run.fold_metrics_csv.exists()
    assert run.predictions_csv.exists()
    assert run.report_md.exists()
    assert len(run.figure_paths) == 4
    assert all(path.stat().st_size > 0 for path in run.figure_paths)
    assert len(run.model_paths) == 8

    rows = _read_csv(run.predictions_csv)
    assert len(rows) == (62 + 42) * 4
    counts = Counter((row["target"], row["model"]) for row in rows)
    assert set(counts.values()) == {62, 42}
    identities = Counter(
        (row["target"], row["model"], row["sample_id"]) for row in rows
    )
    assert set(identities.values()) == {1}
    assert all(
        float(row["prediction_lower"]) <= float(row["prediction_upper"]) for row in rows
    )
    assert all(int(row["calibration_sample_count"]) > 0 for row in rows)

    comparison = _read_csv(run.comparison_csv)
    assert len(comparison) == 8
    assert {row["rank_by_mae"] for row in comparison} == {"1", "2", "3", "4"}
    assert run.summary["evaluation_strategy"] == "leave_one_material_type_out"
    assert run.summary["paper_comparison"]["status"] == "not_directly_comparable"

    with run.model_paths[0].open("rb") as handle:
        model_bundle = pickle.load(handle)
    assert model_bundle["feature_columns"] == [
        "cnt_volume_fraction",
        "interlayer_volume_fraction",
        "glass_transition_temperature_c",
        "manufacturing_pressure_psi",
    ]
    assert hasattr(model_bundle["estimator"], "predict")

    manifest = json.loads(run.manifest_json.read_text(encoding="utf-8"))
    artifact_paths = {entry["path"] for entry in manifest["artifacts"]}
    assert "model_comparison.csv" in artifact_paths
    assert "models/flexural_strength_mpa/random_forest.pkl" in artifact_paths


def test_split_contract_rejects_held_out_sample_in_training(tmp_path: Path) -> None:
    dataset_dir = _prepare_dataset(tmp_path / "source")
    split_path = dataset_dir / "derived" / "grouped_splits.json"
    splits = json.loads(split_path.read_text(encoding="utf-8"))
    fold = splits["targets"]["flexural_strength_mpa"]["folds"][0]
    fold["train_sample_ids"].append(fold["test_sample_ids"][0])
    fold["train_count"] += 1
    split_path.write_text(
        json.dumps(splits, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    _refresh_manifest_artifact(dataset_dir, "grouped_splits")

    bundle = load_cfrp_baseline_dataset(dataset_dir)
    with pytest.raises(ValueError, match="overlapping train and test"):
        validate_target_split_contract(bundle, "flexural_strength_mpa")


def test_outer_test_truth_never_changes_its_prediction_or_interval(
    tmp_path: Path,
) -> None:
    original_dir = _prepare_dataset(tmp_path / "source")
    changed_dir = tmp_path / "changed_dataset"
    shutil.copytree(original_dir, changed_dir)

    original_run = train_cfrp_grouped_baselines(
        original_dir,
        output_root=tmp_path / "original_runs",
        targets=("flexural_strength_mpa",),
        models=("mean",),
        rf_estimators=10,
    )

    normalized = changed_dir / "derived" / "cfrp_experimental_normalized.csv"
    rows = _read_csv(normalized)
    fieldnames = list(rows[0])
    for row in rows:
        if row["material_type_id"] == "1":
            row["flexural_strength_mpa"] = str(
                float(row["flexural_strength_mpa"]) + 10000.0
            )
    with normalized.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    _refresh_manifest_artifact(changed_dir, "normalized_csv")

    changed_run = train_cfrp_grouped_baselines(
        changed_dir,
        output_root=tmp_path / "changed_runs",
        targets=("flexural_strength_mpa",),
        models=("mean",),
        rf_estimators=10,
    )

    def held_out_type_one(run_path: Path) -> dict[str, tuple[float, float, float]]:
        return {
            row["sample_id"]: (
                float(row["prediction"]),
                float(row["prediction_lower"]),
                float(row["prediction_upper"]),
            )
            for row in _read_csv(run_path)
            if row["fold_id"] == "holdout_type_01"
        }

    assert held_out_type_one(original_run.predictions_csv) == held_out_type_one(
        changed_run.predictions_csv
    )


def test_baseline_cli_emits_structured_success_and_error(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    dataset_dir = _prepare_dataset(tmp_path / "source")
    exit_code = baseline_cli_main(
        [
            "--dataset-dir",
            str(dataset_dir),
            "--output-root",
            str(tmp_path / "runs"),
            "--target",
            "flexural_strength_mpa",
            "--model",
            "mean",
            "--rf-estimators",
            "10",
        ]
    )
    captured = capsys.readouterr()
    assert exit_code == 0
    assert captured.err == ""
    payload = json.loads(captured.out)
    assert payload["status"] == "completed_with_warnings"
    assert payload["best_models_by_grouped_mae"] == {"flexural_strength_mpa": "mean"}

    exit_code = baseline_cli_main(
        ["--dataset-dir", str(tmp_path / "missing"), "--model", "mean"]
    )
    captured = capsys.readouterr()
    assert exit_code == 2
    assert captured.out == ""
    assert json.loads(captured.err)["status"] == "error"


def test_validation_audit_compares_protocols_and_flags_duplicate_leakage(
    tmp_path: Path,
) -> None:
    dataset_dir = _prepare_dataset(tmp_path / "source")
    _make_exact_duplicate(
        dataset_dir,
        retained_sample_id="cfrp_type_01_sample_01",
        duplicate_sample_id="cfrp_type_01_sample_02",
    )

    run = run_cfrp_validation_audit(
        dataset_dir,
        output_root=tmp_path / "audits",
        targets=("flexural_strength_mpa",),
        models=("mean", "ridge"),
        random_state=7,
        rf_estimators=10,
    )

    assert run.manifest_json.exists()
    assert run.summary_json.exists()
    assert run.comparison_csv.exists()
    assert run.predictions_csv.exists()
    assert run.duplicate_clusters_csv.exists()
    assert run.report_md.exists()
    assert len(run.figure_paths) == 1
    assert run.figure_paths[0].stat().st_size > 0

    duplicates = _read_csv(run.duplicate_clusters_csv)
    assert len(duplicates) == 1
    assert duplicates[0]["record_count"] == "2"
    assert duplicates[0]["duplicate_extra_count"] == "1"
    assert duplicates[0]["retained_sample_id"] == "cfrp_type_01_sample_01"
    assert duplicates[0]["removed_sample_ids"] == "cfrp_type_01_sample_02"

    comparison = _read_csv(run.comparison_csv)
    assert len(comparison) == 8
    assert {row["protocol"] for row in comparison} == {
        "grouped_raw",
        "grouped_deduplicated",
        "row_loocv_raw",
        "row_loocv_deduplicated",
    }
    assert {
        int(row["sample_count"])
        for row in comparison
        if row["protocol"].endswith("deduplicated")
    } == {61}
    assert {
        int(row["duplicate_leakage_sample_count"])
        for row in comparison
        if row["protocol"] == "row_loocv_raw"
    } == {2}
    assert {
        int(row["duplicate_leakage_sample_count"])
        for row in comparison
        if row["protocol"] != "row_loocv_raw"
    } == {0}

    predictions = _read_csv(run.predictions_csv)
    assert len(predictions) == 492
    leaking_ids = {
        row["sample_id"]
        for row in predictions
        if row["protocol"] == "row_loocv_raw"
        and row["model"] == "ridge"
        and row["exact_duplicate_in_training"] == "True"
    }
    assert leaking_ids == {
        "cfrp_type_01_sample_01",
        "cfrp_type_01_sample_02",
    }
    assert not any(
        row["exact_duplicate_in_training"] == "True"
        for row in predictions
        if row["protocol"] == "grouped_raw"
    )

    assert run.summary["release_gate_protocol"] == "grouped_raw"
    assert run.summary["duplicate_audit"]["duplicate_extra_rows"] == 1
    assert run.summary["paper_comparison"]["status"] == "not_directly_comparable"


def test_validation_audit_cli_emits_structured_success_and_error(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    dataset_dir = _prepare_dataset(tmp_path / "source")
    exit_code = validation_audit_cli_main(
        [
            "--dataset-dir",
            str(dataset_dir),
            "--output-root",
            str(tmp_path / "audits"),
            "--target",
            "flexural_strength_mpa",
            "--model",
            "mean",
            "--rf-estimators",
            "10",
        ]
    )
    captured = capsys.readouterr()
    assert exit_code == 0
    assert captured.err == ""
    payload = json.loads(captured.out)
    assert payload["status"] == "completed_with_warnings"
    assert payload["release_gate_protocol"] == "grouped_raw"

    exit_code = validation_audit_cli_main(
        ["--dataset-dir", str(tmp_path / "missing"), "--model", "mean"]
    )
    captured = capsys.readouterr()
    assert exit_code == 2
    assert captured.out == ""
    assert json.loads(captured.err)["status"] == "error"
